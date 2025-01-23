import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox, Progressbar
import pandas as pd
import os
import datetime
import threading
from queue import Queue
import xlsxwriter
import re
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font
import gc
import psutil
import time
from functools import lru_cache
from concurrent.futures import ThreadPoolExecutor
from queue import Queue
import threading


class DuplicateFinderApp:
    def __init__(self, root):
        self.root = root
        self.setup_logging()
        self.setup_memory_monitor()
        self.initialize_gui()
        self.input_queue = Queue()
        self.processing_queue = Queue()
        self.output_queue = Queue()
        self.max_workers = min(32, (os.cpu_count() or 1) + 4)
        # Add case_sensitive flag, default to False for case-insensitive matching
        self.case_sensitive = False

    def setup_logging(self):
        # Create logs directory if it doesn't exist
        log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
        os.makedirs(log_dir, exist_ok=True)
        
        # Create timestamped log file
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = os.path.join(log_dir, f'duplicate_finder_{timestamp}.log')
        
        # Configure logging with more detailed format
        logging.basicConfig(
            filename=log_file,
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
            filemode='w'
        )
        self.logger = logging.getLogger(__name__)
        self.log_file = log_file

    def log_error(self, error_type, file_name, error_msg):
        error_detail = f"{error_type} - File: {file_name} - {error_msg}"
        self.logger.error(error_detail)
        
        # Also save to Excel if error list exists
        if hasattr(self, 'error_files'):
            self.error_files.append({
                'Timestamp': datetime.datetime.now(),
                'Error Type': error_type,
                'File': file_name,
                'Error': error_msg
            })

    def setup_memory_monitor(self):
        self.memory_threshold = 85  # Percentage
        self.process = psutil.Process()

    def check_memory_usage(self):
        memory_percent = self.process.memory_percent()
        if memory_percent > self.memory_threshold:
            self.logger.warning(f"High memory usage: {memory_percent:.2f}%")
            return False
        return True

    @lru_cache(maxsize=128)
    def get_sheet_names(self, file_path):
        try:
            with pd.ExcelFile(file_path) as xls:
                return xls.sheet_names
        except Exception as e:
            self.logger.error(f"Error reading sheet names from {file_path}: {str(e)}")
            raise

    def process_excel_chunk(self, file_path, sheet_name, chunk_size=1000):
        try:
            chunks = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                dtype=str,
                chunksize=chunk_size,
                usecols=lambda x: any(x.lower().contains(key) for key in ['barcode', 'code', 'id'])
            )
            return pd.concat([self.process_chunk(chunk) for chunk in chunks])
        except Exception as e:
            self.logger.error(f"Error processing file {file_path}: {str(e)}")
            raise

    def process_chunk(self, chunk):
        if not self.check_memory_usage():
            self.clear_memory()
        return self.find_barcodes_in_dataframe(chunk)

    def clear_memory(self):
        gc.collect()
        if hasattr(self, 'current_df'):
            del self.current_df
        if hasattr(self, 'processed_data'):
            del self.processed_data

    def retry_operation(self, operation, max_retries=3):
        for attempt in range(max_retries):
            try:
                return operation()
            except Exception as e:
                self.logger.warning(f"Attempt {attempt + 1} failed: {str(e)}")
                if attempt == max_retries - 1:
                    raise
                time.sleep(1)

    def process_files(self):
        try:
            self.disable_controls()
            self.clear_memory()

            results = []
            for file in self.selected_files:
                if not self.check_memory_usage():
                    raise MemoryError("Insufficient memory to continue processing")

                result = self.retry_operation(
                    lambda: self.process_single_file(file)
                )
                results.append(result)

            self.generate_report(results)

        except Exception as e:
            self.logger.error(f"Critical error in process_files: {str(e)}")
            self.queue.put(("error", str(e)))
        finally:
            self.enable_controls()
            self.clear_memory()

    def save_report(self, df, filename):
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
                self.apply_formatting(writer)
            self.show_success_with_open("Process completed successfully!", filename)
        except Exception as e:
            self.logger.error(f"Error saving report: {str(e)}")
            raise

    def show_success_with_open(self, message, filepath):
        dialog = tk.Toplevel()
        dialog.title("Success")
        dialog.geometry("300x150")
        dialog.grab_set()  # Make dialog modal

        # Center the dialog
        dialog.geometry("+%d+%d" % (dialog.winfo_screenwidth() / 2 - 150,
                                    dialog.winfo_screenheight() / 2 - 75))

        # Message
        label = tk.Label(dialog, text=message, wraplength=250, pady=20)
        label.pack()

        # Buttons frame
        button_frame = tk.Frame(dialog)
        button_frame.pack(pady=10)

        def open_file():
            os.startfile(filepath)
            dialog.destroy()

        # Open button
        open_btn = tk.Button(button_frame, text="Open", command=open_file)
        open_btn.pack(side=tk.LEFT, padx=5)

        # OK button
        ok_btn = tk.Button(button_frame, text="OK", command=dialog.destroy)
        ok_btn.pack(side=tk.LEFT, padx=5)

        dialog.wait_window()

    def initialize_gui(self):
        self.root.title("ICON Barcode Duplicate Finder v2.3.4")
        self.root.geometry("600x500")

        # Custom ICON barcode patterns
        self.barcode_patterns = {
            'ICON-17': r'^ICON\d{13}$',  # ICON followed by 13 digits
            'ICON-18': r'^ICON\d{3}[A-Z]\d{10}$',  # ICON + 3 digits + 1 letter + 10 digits
            'ICON-20': r'^ICON\d{5}[A-Z]\d{10}$'  # ICON + 5 digits + 1 letter + 10 digits
        }

        self.selected_files = []
        self.sheet_selection_comboboxes = []
        self.sheet_headers = {}

        # Queue for thread communication
        self.queue = Queue()

        # Create GUI elements
        self.create_gui()

    def reset_selection(self):
        self.selected_files = []
        self.file_label.config(text="No files selected")
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.sheet_selection_comboboxes = []
        self.sheet_headers = {}

        # Reset the progress bar and status
        self.progress["value"] = 0  # Reset the progress bar to 0
        self.status_var.set("")  # Clear the status label

    def disable_controls(self):
        self.file_button.config(state="disabled")
        self.folder_button.config(state="disabled")
        self.start_button.config(state="disabled")
        self.reset_button.config(state="disabled")
        for combobox in self.sheet_selection_comboboxes:
            combobox.config(state="disabled")

    def enable_controls(self):
        self.file_button.config(state="normal")
        self.folder_button.config(state="normal")
        self.start_button.config(state="normal")
        self.reset_button.config(state="normal")
        for combobox in self.sheet_selection_comboboxes:
            combobox.config(state="readonly")

    def start_processing(self):
        if not self.selected_files:
            messagebox.showwarning("Warning", "Please select files first.")
            return

        self.disable_controls()
        thread = threading.Thread(target=self.process_files, daemon=True)
        thread.start()
        self.check_queue()

    def create_gui(self):
        # Create a frame for the buttons
        button_frame = tk.Frame(self.root)
        button_frame.pack(pady=10)

        # File selection button
        self.file_button = tk.Button(button_frame, text="Select Files", command=self.select_files)
        self.file_button.pack(side="left", padx=5)

        # Folder selection button
        self.folder_button = tk.Button(button_frame, text="Select Folder", command=self.select_folder)
        self.folder_button.pack(side="left", padx=5)

        self.file_label = tk.Label(self.root, text="No files selected")
        self.file_label.pack()

        # Create a scrollable canvas
        self.canvas_frame = tk.Frame(self.root)
        self.canvas_frame.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(self.canvas_frame)
        self.scrollbar = tk.Scrollbar(self.canvas_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Bind scrolling events
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Up>", lambda e: self.canvas.yview_scroll(-1, "units"))
        self.canvas.bind_all("<Down>", lambda e: self.canvas.yview_scroll(1, "units"))
        self.canvas.bind_all("<Prior>", lambda e: self.canvas.yview_scroll(-1, "pages"))
        self.canvas.bind_all("<Next>", lambda e: self.canvas.yview_scroll(1, "pages"))

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # Progress frame
        self.progress_frame = tk.Frame(self.root)
        self.progress_frame.pack(fill="x", pady=10)

        self.progress = Progressbar(
            self.progress_frame,
            orient="horizontal",
            mode="determinate",
            length=300
        )
        self.progress.pack(pady=5)

        self.status_var = tk.StringVar()
        self.status_label = tk.Label(
            self.progress_frame,
            textvariable=self.status_var,
            wraplength=500,
            height=2,
            justify=tk.CENTER
        )
        self.status_label.pack(fill="x", padx=10)

        # Start button
        self.start_button = tk.Button(
            self.root,
            text="Start Duplicate Check",
            command=self.start_processing
        )
        self.start_button.pack(pady=10)

        # Reset button (add this after the Start button)
        self.reset_button = tk.Button(
            self.root,
            text="Reset Selection",
            command=self.reset_selection
        )
        self.reset_button.pack(pady=5)

    def detect_barcodes(self, value):
        """
        Detect if a value matches ICON barcode pattern
        Returns (is_barcode, barcode_type)
        """
        if pd.isna(value):
            return False, None

        # Convert to string and remove any whitespace
        str_value = str(value).strip()

        # Skip empty strings
        if not str_value:
            return False, None

        # Check for exact length (17 or 18 characters)
        if len(str_value) not in [17, 18, 20]:
            return False, None

        # Check if starts with 'ICON'
        if not str_value.startswith('ICON'):
            return False, None

        # For 17-character format
        if len(str_value) == 17:
            if re.match(self.barcode_patterns['ICON-17'], str_value):
                return True, 'ICON-17'

        # For 18-character format
        elif len(str_value) == 18:
            if re.match(self.barcode_patterns['ICON-18'], str_value):
                return True, 'ICON-18'

        # For 20-character format
        elif len(str_value) == 20:
            if re.match(self.barcode_patterns['ICON-20'], str_value):
                return True, 'ICON-20'

        return False, None

    def find_barcodes_in_dataframe(self, df):
        """Find all ICON barcode values in a DataFrame."""
        barcodes = []

        for column in df.columns:
            for idx, value in enumerate(df[column]):
                is_barcode, barcode_type = self.detect_barcodes(value)
                if is_barcode:
                    barcodes.append({
                        'value': str(value).strip(),
                        'column': column,
                        'row': idx + 2,  # Adding 2 for Excel row number (1-based + header)
                        'type': barcode_type
                    })

        return barcodes

    def _on_mousewheel(self, event):
        if event.num == 5 or event.delta < 0:
            self.canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            self.canvas.yview_scroll(-1, "units")

    def is_valid_excel_file(self, filename):
        """Check if the file is a valid Excel file (not temporary and has correct extension)"""
        base_name = os.path.basename(filename)
        return (
            not base_name.startswith("~$") and  # Skip temporary files
            base_name.endswith((".xlsx", ".xls", ".xlsm"))  # Must be Excel file
        )

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="Select Excel Files",
            filetypes=[("Excel Files", "*.xlsx *.xls *.xlsm")]
        )
        if files:
            # Convert selected files to absolute paths
            selected_absolute_paths = {os.path.abspath(f) for f in files}

            # Filter out temporary files
            valid_files = {f for f in selected_absolute_paths if self.is_valid_excel_file(f)}

            if not valid_files:
                messagebox.showwarning("Warning", "No valid Excel files selected. Temporary files (~$) will be skipped.")
                return

            # Convert existing files to absolute paths and combine with new unique files
            existing_absolute_paths = {os.path.abspath(f) for f in self.selected_files}
            self.selected_files = list(existing_absolute_paths.union(valid_files))

            skipped = len(files) - len(valid_files)
            if skipped > 0:
                self.file_label.config(
                    text=f"{len(valid_files)} file(s) selected ({skipped} temporary file(s) skipped)")
            else:
                self.file_label.config(text=f"{len(self.selected_files)} file(s) selected")
            self.display_file_selection()
        else:
            self.file_label.config(text="No files selected")

    def is_valid_excel_file(self, filename):
        # Check if the file is a valid Excel file (not temporary and has correct extension)
        base_name = os.path.basename(filename)
        return (
            not base_name.startswith("~$") and  # Skip temporary files
            base_name.endswith((".xlsx", ".xls", ".xlsm"))  # Added .xlsm
        )

    def get_excel_engine(self, file_path):
        """Determine the appropriate engine based on file extension"""
        file_extension = os.path.splitext(file_path)[1].lower()
        if file_extension == '.xls':
            return 'xlrd'
        else:  # .xlsx and .xlsm files
            return 'openpyxl'

    def select_folder(self):
        folder_selected = filedialog.askdirectory(title="Select Folder")

        if not folder_selected:
            return

        excel_files = set()
        skipped_files = 0
        # Walk through all subdirectories and files
        for root, dirs, files in os.walk(folder_selected):
            for file in files:
                full_path = os.path.abspath(os.path.join(root, file))
                if self.is_valid_excel_file(full_path):
                    excel_files.add(full_path)
                elif file.endswith(('.xlsx', '.xls', '.xlsm')):  # Count skipped Excel files
                    skipped_files += 1

        if excel_files:
            # Convert existing files to absolute paths and combine with new unique files
            existing_absolute_paths = {os.path.abspath(f) for f in self.selected_files}
            self.selected_files = list(existing_absolute_paths.union(excel_files))

            if skipped_files > 0:
                self.file_label.config(
                    text=f"{len(excel_files)} file(s) selected ({skipped_files} temporary file(s) skipped)")
            else:
                self.file_label.config(text=f"{len(self.selected_files)} file(s) selected")

            self.display_file_selection()
        else:
            if skipped_files > 0:
                messagebox.showinfo("Information",
                                    "Only temporary Excel files were found. These files are skipped.")
            else:
                messagebox.showinfo("Information", "No Excel files found in the selected folder and its subfolders.")

    def display_file_selection(self):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.sheet_selection_comboboxes = []
        self.sheet_headers = {}

        for idx, file in enumerate(self.selected_files):
            try:
                # Get appropriate engine for the file type
                engine = self.get_excel_engine(file)

                # Read Excel file with appropriate engine
                excel_file = pd.ExcelFile(file, engine=engine)
                sheet_names = excel_file.sheet_names
                self.sheet_headers[file] = sheet_names

                file_frame = tk.Frame(self.scrollable_frame)
                file_frame.pack(fill="x", pady=5)

                sheet_label = tk.Label(
                    file_frame,
                    text=f"File {idx + 1}: {os.path.basename(file)}",
                    width=50,
                    anchor="w"
                )
                sheet_label.pack(side="left", padx=10)

                sheet_combobox = Combobox(
                    file_frame,
                    values=sheet_names,
                    state="readonly",
                    width=17
                )
                sheet_combobox.pack(side="left", padx=10)
                sheet_combobox.current(0)
                self.sheet_selection_comboboxes.append(sheet_combobox)
            except Exception as e:
                messagebox.showerror("Error", f"Error reading file {os.path.basename(file)}: {str(e)}")

    def process_files(self):
        try:
            all_barcodes = []
            file_summary = []
            error_files = []
            file_paths_dict = {}  # Dictionary to store file paths with filenames as keys
            total_files = len(self.selected_files)

            # Calculate progress weights (80% for file processing, 10% for duplicates, 10% for report)
            file_progress_weight = 80
            progress_per_file = file_progress_weight / total_files if total_files > 0 else 0

            for idx, file in enumerate(self.selected_files):
                try:
                    selected_sheet = self.sheet_selection_comboboxes[idx].get()
                    file_name = os.path.basename(file)
                    file_path = os.path.abspath(file)
                    file_paths_dict[file_name] = file_path  # Store file path with filename as key

                    current_progress = idx * progress_per_file
                    self.update_status(current_progress, f"Reading {file_name}...")

                    engine = self.get_excel_engine(file)
                    df = pd.read_excel(file, sheet_name=selected_sheet, dtype=str, engine=engine)
                    self.update_status(current_progress + (progress_per_file / 2), f"Scanning for ICON barcodes in {file_name}...")

                    barcodes = self.find_barcodes_in_dataframe(df)

                    file_summary.append({
                        'FILE_NAME': file_name,
                        'BARCODE_COUNT': len(barcodes),
                        'PATH': file_path,
                        'STATUS': 'Processed successfully'
                    })

                    if not barcodes:
                        self.update_status(current_progress + progress_per_file,
                                           f"No ICON barcodes found in {file_name}, continuing...")
                        continue

                    for barcode in barcodes:
                        all_barcodes.append({
                            'BARCODE': barcode['value'],
                            'FILE_NAME': file_name,
                            'FORMAT': barcode['type'],
                            'FILE_PATH': file_path  # Add file path to barcode data
                        })

                except Exception as e:
                    error_message = f"Error processing {file_name}: {str(e)}"
                    error_files.append(error_message)
                    file_summary.append({
                        'FILE_NAME': file_name,
                        'BARCODE_COUNT': 0,
                        'PATH': file_path,
                        'STATUS': f'Failed: {str(e)}'
                    })
                    self.update_status(current_progress + progress_per_file, f"Skipping {file_name} due to error...")
                    continue

            #Update the no barcodes case to include filename parameter
            if not all_barcodes and not error_files:
                self.queue.put(("complete", False, "No ICON barcodes found in any of the selected files.", None))
                self.update_status(0, "")
                return

            self.update_status(90, "Processing duplicates...")

            file_summary_df = pd.DataFrame(file_summary)
            file_summary_df = file_summary_df.sort_values('BARCODE_COUNT', ascending=False)

            if all_barcodes:
                barcode_df = pd.DataFrame(all_barcodes)
                duplicate_barcodes = barcode_df[barcode_df.duplicated("BARCODE", keep=False)]

                if duplicate_barcodes.empty:
                    success_msg = "No duplicate ICON barcodes found."
                    if error_files:
                        success_msg += f"\n\nWarning: {len(error_files)} file(s) were skipped due to errors."
                    self.update_status(100, "Complete")
                    self.queue.put(("complete", True, success_msg, None))
                    return

                self.update_status(95, "Compiling Duplicates...")

                grouped_duplicates = []
                for barcode, group in duplicate_barcodes.groupby("BARCODE"):
                    copies = len(group)
                    row_data = [barcode, copies]
                    for _, row in group.iterrows():
                        row_data.append((row['FILE_NAME'], row['FILE_PATH']))
                    grouped_duplicates.append(row_data)

                max_files = max(len(row) - 2 for row in grouped_duplicates)
                headers = ["DUPLICATE_BARCODES", "COPIES"] + [f"FILE_NAME{i + 1}" for i in range(max_files)]

                # Prepare data for DataFrame (separate names and paths)
                aligned_duplicates = []
                for row in grouped_duplicates:
                    new_row = [row[0], row[1]]  # Barcode and copies
                    file_tuples = row[2:]  # List of (name, path) tuples
                    # Add file names only (paths will be used later for hyperlinks)
                    new_row.extend([tup[0] if isinstance(tup, tuple) else "" for tup in file_tuples + [("", "")] * (max_files - len(file_tuples))])
                    aligned_duplicates.append(new_row)

                duplicates_df = pd.DataFrame(aligned_duplicates, columns=headers)

                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                destination_path = os.path.expanduser("~")
                folder_path = os.path.join(destination_path, "Desktop", "DUPLICATE_BARCODES")
                os.makedirs(folder_path, exist_ok=True)
                output_filename = os.path.join(folder_path, f"ICON_Duplicates_{timestamp}.xlsx")

                # Save the basic structure first
                with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                    duplicates_df.to_excel(writer, sheet_name='Detailed_Report', index=False)
                    file_summary_df.to_excel(writer, sheet_name='File_Summary', index=False)

                    summary_data = {
                        'Metric': [
                            'Total Files Processed',
                            'Successfully Processed Files',
                            'Failed Files',
                            'Total ICON Barcodes Found',
                            'Unique Barcodes',
                            'Duplicate Barcodes',
                            '17-Character Barcodes',
                            '18-Character Barcodes',
                            '20-Character Barcodes'
                        ],
                        'Value': [
                            len(self.selected_files),
                            len(file_summary_df[file_summary_df['STATUS'].str.startswith('Processed')]),
                            len(error_files),
                            len(barcode_df),
                            len(barcode_df['BARCODE'].unique()),
                            len(duplicate_barcodes['BARCODE'].unique()),
                            len(barcode_df[barcode_df['FORMAT'] == 'ICON-17']),
                            len(barcode_df[barcode_df['FORMAT'] == 'ICON-18']),
                            len(barcode_df[barcode_df['FORMAT'] == 'ICON-20'])
                        ]
                    }
                    pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

                # Now add hyperlinks using openpyxl
                wb = load_workbook(output_filename)

                # Add hyperlinks to Detailed_Report sheet
                self.update_status(97, "Adding Hyperlink on Detailed_Repor...")
                ws_detailed = wb['Detailed_Report']
                for row in range(2, ws_detailed.max_row + 1):  # Start from row 2 to skip header
                    for col in range(3, ws_detailed.max_column + 1):  # Start from column 3 (FILE_NAME1)
                        cell = ws_detailed.cell(row=row, column=col)
                        if cell.value:  # If there's a filename
                            file_path = file_paths_dict.get(cell.value)  # Use dictionary get() method
                            if file_path:
                                cell.hyperlink = file_path
                                cell.font = Font(color="0000FF", underline="single")  # Blue, underlined

                # Add hyperlinks to File_Summary sheet
                self.update_status(98, "Adding Hyperlink on File_Summary...")
                ws_summary = wb['File_Summary']
                path_col = None
                # Find the PATH column
                for col in range(1, ws_summary.max_column + 1):
                    if ws_summary.cell(row=1, column=col).value == 'PATH':
                        path_col = col
                        break

                if path_col:
                    for row in range(2, ws_summary.max_row + 1):
                        cell = ws_summary.cell(row=row, column=path_col)
                        if cell.value:
                            cell.hyperlink = cell.value
                            cell.font = Font(color="0000FF", underline="single")

                # Save the workbook with hyperlinks
                wb.save(output_filename)

                success_msg = f"Found {len(duplicate_barcodes['BARCODE'].unique())} duplicate ICON barcodes. "
                if error_files:
                    success_msg += f"\n\nWarning: {len(error_files)} file(s) were skipped due to errors. "
                success_msg += f"\nReport saved to '{output_filename}'"

                self.update_status(100, "Saved.")
                self.queue.put(("complete", True, success_msg, output_filename))
                self.root.after(1000, lambda: self.update_status(0, ""))

            else:
                msg = "No valid barcodes found in processable files."
                if error_files:
                    msg += f"\n\nWarning: {len(error_files)} file(s) were skipped due to errors:"
                    for error in error_files:
                        msg += f"\n- {error}"
                self.update_status(100, "")
                self.queue.put(("complete", True, msg, None))  # Add None as filename
                self.root.after(1000, lambda: self.update_status(0, ""))

            # Modify existing error handling code:
            if error_files:
                for error in error_files:
                    self.log_error("Processing Error", "Unknown", error)
                msg += f"\n\nErrors have been logged to: {self.log_file}"

        except Exception as e:
            self.queue.put(("complete", False, f"A critical error occurred: {str(e)}", None))
            self.root.after(1000, lambda: self.update_status(0, ""))

    def update_status(self, progress, status):
        self.queue.put(("status", progress, status))

    #Add the check_queue method with proper error handling
    def check_queue(self):
        """Process messages from the queue"""
        try:
            while self.queue.qsize():
                msg = self.queue.get(0)
                msg_type = msg[0]

                if msg_type == "status":
                    progress, status = msg[1:]
                    self.progress["value"] = progress
                    self.status_var.set(status)
                elif msg_type == "complete":
                    try:
                        _, success, message, filename = msg
                    except ValueError:
                        # Handle case where filename is missing
                        _, success, message = msg
                        filename = None
                    
                    if success and filename:
                        self.show_success_with_open(message, filename)
                    else:
                        messagebox.showinfo("Process Complete", message)
                    
                    self.enable_controls()
                elif msg_type == "error":
                    messagebox.showerror("Error", msg[1])
                    self.enable_controls()

        except Exception as e:
            self.logger.error(f"Error in check_queue: {str(e)}")
            messagebox.showerror("Error", "An error occurred while processing results")
            self.enable_controls()
        
        finally:
            self.root.after(100, self.check_queue)

    def create_excel_with_hyperlinks(self, df: pd.DataFrame, output_path: str, chunk_size: int = 5000) -> None:
        try:
            # Phase 1: Fast initial write with xlsxwriter
            with xlsxwriter.Workbook(output_path, {'constant_memory': True}) as workbook:
                worksheet = workbook.add_worksheet('Duplicates')
                
                # Write headers
                for col, header in enumerate(df.columns):
                    worksheet.write(0, col, header)
                
                # Write data in chunks
                total_rows = len(df)
                for start_idx in range(0, total_rows, chunk_size):
                    end_idx = min(start_idx + chunk_size, total_rows)
                    chunk = df.iloc[start_idx:end_idx]
                    
                    for row_idx, row in enumerate(chunk.values, start_idx + 1):
                        for col_idx, value in enumerate(row):
                            worksheet.write(row_idx, col_idx, value)
                    
                    del chunk
                    gc.collect()
                    self.update_progress((start_idx + chunk_size) / total_rows * 50)  # First 50%
            
            # Phase 2: Add hyperlinks with openpyxl
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Process hyperlinks in chunks
            hyperlink_columns = [idx for idx, col in enumerate(df.columns) if 'path' in col.lower()]
            total_links = len(hyperlink_columns) * total_rows
            links_processed = 0
            
            for row_idx in range(2, total_rows + 2):  # Skip header
                for col_idx in hyperlink_columns:
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    if cell.value and isinstance(cell.value, str):
                        cell.hyperlink = cell.value
                        cell.font = Font(color="0000FF", underline="single")
                    
                    links_processed += 1
                    if links_processed % 100 == 0:
                        self.update_progress(50 + (links_processed / total_links * 50))  # Last 50%
            
            wb.save(output_path)
            wb.close()
            
            self.logger.info(f"Excel file created successfully: {output_path}")
            self.clear_memory()
            
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {str(e)}")
            raise

    def update_progress(self, percentage: float) -> None:
        if hasattr(self, 'progress_var'):
            self.progress_var.set(percentage)
            self.root.update_idletasks()

    def clear_memory(self) -> None:
        gc.collect()
        if hasattr(self, 'df'):
            del self.df
        self.logger.info(f"Current memory usage: {self.process.memory_percent()}%")

    def process_files_threaded(self, grouped_duplicates):
        # Initialize thread pools
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Start writer thread
            writer_thread = threading.Thread(
                target=self.excel_writer_thread,
                args=(self.output_queue,)
            )
            writer_thread.start()

            # Submit processing tasks
            chunk_size = 1000
            total_rows = len(grouped_duplicates)
            futures = []

            for i in range(0, total_rows, chunk_size):
                chunk = grouped_duplicates[i:i + chunk_size]
                futures.append(
                    executor.submit(self.process_chunk, chunk)
                )

            # Collect results
            for future in futures:
                self.output_queue.put(future.result())

            # Signal completion
            self.output_queue.put(None)
            writer_thread.join()

    def process_chunk(self, chunk):
        processed_rows = []
        for row in chunk:
            new_row = [row[0], row[1]]  # Barcode and copies
            file_tuples = row[2:]
            new_row.extend([tup[0] if isinstance(tup, tuple) else "" 
                          for tup in file_tuples])
            processed_rows.append(new_row)
        return processed_rows

    def excel_writer_thread(self, output_queue):
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        destination_path = os.path.expanduser("~")
        folder_path = os.path.join(destination_path, "Desktop", "DUPLICATE_BARCODES")
        os.makedirs(folder_path, exist_ok=True)
        output_filename = os.path.join(folder_path, f"ICON_Duplicates_{timestamp}.xlsx")

        with xlsxwriter.Workbook(output_filename, {'constant_memory': True}) as workbook:
            worksheet = workbook.add_worksheet('Detailed_Report')
            row_num = 0

            while True:
                chunk = output_queue.get()
                if chunk is None:  # Exit signal
                    break

                for row in chunk:
                    for col, value in enumerate(row):
                        worksheet.write(row_num, col, value)
                    row_num += 1

                self.update_progress(row_num)

    def compare_patterns(self, str1, str2):
        """Compare two strings case-insensitively"""
        if self.case_sensitive:
            return str1 == str2
        return str1.lower() == str2.lower()

    def find_duplicates(self, data):
        # Convert patterns to lowercase for comparison while keeping original data
        lowercase_data = {k: v.lower() if isinstance(v, str) else v 
                         for k, v in data.items()}
        duplicates = []
        for i in range(len(data)):
            for j in range(i + 1, len(data)):
                if self.compare_patterns(lowercase_data[i], lowercase_data[j]):
                    duplicates.append((i, j))
        return duplicates

def open_file(filepath):
    os.startfile(filepath)

if __name__ == "__main__":
    root = tk.Tk()
    app = DuplicateFinderApp(root)
    root.mainloop()
